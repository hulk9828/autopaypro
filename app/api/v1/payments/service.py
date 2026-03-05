import io
from datetime import date, datetime, timedelta
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy import case, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.payments.schemas import (
    DueCustomerItem,
    DueEntryItem,
    DueInstallmentItem,
    NOTIFICATION_TYPE_DISPLAY,
    NotificationItem,
    OverdueItem,
    TransactionItem,
)
from app.core.config import settings
from app.core.email import send_overdue_reminder_email
from app.core.exceptions import AppException
from app.core.notification_service import scope_key_for_payment, send_payment_notification
from app.core.stripe_client import create_payment_intent, confirm_payment_intent_with_token, get_payment_intent
from app.api.v1.notifications.service import send_notification_to_customers
from app.models.customer import Customer
from app.models.loan import Loan
from app.models.payment import Payment, PaymentMode, PaymentStatus
from app.models.payment_notification_log import PaymentNotificationLog
from app.models.vehicle import Vehicle
from app.core.utils import ensure_non_negative_amount as _ensure_non_negative_amount
from app.core.loan_schedule import get_due_dates_range as _schedule_due_dates_range


def _loan_status_display(loan: Loan | None) -> str:
    """Return 'completed' if loan is closed (fully paid), else 'active'. For API display."""
    if not loan:
        return "active"
    return "completed" if getattr(loan, "status", "active") == "closed" else "active"


def _get_due_dates_range(loan: Loan, from_date: date, to_date: date) -> list[datetime]:
    """Return due datetimes for a loan between from_date and to_date (uses loan.lease_payment_type)."""
    payment_type = getattr(loan, "lease_payment_type", "bi_weekly") or "bi_weekly"
    return _schedule_due_dates_range(
        loan.created_at,
        loan.loan_term_months,
        payment_type,
        from_date,
        to_date,
    )


def _parse_due_date_iso(due_date_iso: str) -> datetime:
    """Parse ISO date/datetime string to datetime."""
    s = due_date_iso.strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        raise ValueError("Invalid due_date_iso format")


def _round_amount(val: float) -> float:
    return round(max(0.0, val), 2)


class PaymentService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def _get_paid_amount_per_due_date(self, loan_id: UUID, completed_only: bool = True) -> dict[date, float]:
        """
        Aggregate paid amount per due_date from all payments.
        Legacy: due_date + amount. Flexible: applied_installments JSON.
        """
        q = select(Payment).where(Payment.loan_id == loan_id)
        if completed_only:
            q = q.where(Payment.status == PaymentStatus.completed.value)
        result = await self.db.execute(q)
        paid: dict[date, float] = {}
        for p in result.scalars().all():
            if getattr(p, "applied_installments", None) and isinstance(p.applied_installments, list):
                for entry in p.applied_installments:
                    if isinstance(entry, dict):
                        due_str = entry.get("due_date")
                        amt = float(entry.get("applied_amount") or 0)
                        if due_str and amt > 0:
                            try:
                                due_dt = datetime.fromisoformat(due_str.replace("Z", "+00:00"))
                                d = due_dt.date() if hasattr(due_dt, "date") else due_dt
                                paid[d] = paid.get(d, 0) + amt
                            except (ValueError, TypeError):
                                pass
            else:
                d = p.due_date.date() if hasattr(p.due_date, "date") else p.due_date
                amt = _ensure_non_negative_amount(p.amount)
                paid[d] = paid.get(d, 0) + amt
        return paid

    async def _get_paid_due_dates_for_loan(self, loan_id: UUID) -> set[date]:
        """Set of (due_date as date) fully paid for this loan (paid_amount >= installment_amount)."""
        loan = await self.db.get(Loan, loan_id)
        if not loan:
            return set()
        paid_per = await self._get_paid_amount_per_due_date(loan_id, completed_only=False)
        today = date.today()
        from_date = today - timedelta(days=365)
        to_date = today + timedelta(days=365 * 2)
        due_dates = _get_due_dates_range(loan, from_date, to_date)
        amt_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
        return {d.date() for d in due_dates if paid_per.get(d.date(), 0) >= amt_per - 0.01}

    async def _get_paid_due_dates_for_loan_completed(self, loan_id: UUID) -> set[date]:
        """Set of (due_date as date) fully paid with completed payments only."""
        loan = await self.db.get(Loan, loan_id)
        if not loan:
            return set()
        paid_per = await self._get_paid_amount_per_due_date(loan_id, completed_only=True)
        today = date.today()
        from_date = today - timedelta(days=365)
        to_date = today + timedelta(days=365 * 2)
        due_dates = _get_due_dates_range(loan, from_date, to_date)
        amt_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
        return {d.date() for d in due_dates if paid_per.get(d.date(), 0) >= amt_per - 0.01}

    async def get_paid_due_dates_for_loan(self, loan_id: UUID) -> set[date]:
        """Public: set of due dates fully paid (paid_amount >= installment_amount). For use by CustomerService etc."""
        return await self._get_paid_due_dates_for_loan_completed(loan_id)

    async def get_paid_amount_per_due_date(self, loan_id: UUID) -> dict[date, float]:
        """Public: paid amount per due_date for this loan (completed payments only)."""
        return await self._get_paid_amount_per_due_date(loan_id, completed_only=True)

    async def get_next_unpaid_due(self, loan_id: UUID, customer_id: UUID) -> tuple[datetime, float] | None:
        """
        Get the next unpaid due date and amount for the loan.
        Returns (due_datetime, amount) or None if no unpaid due (e.g. all paid, loan closed, or invalid).
        """
        loan = await self.db.get(Loan, loan_id)
        if not loan or loan.customer_id != customer_id:
            return None
        if getattr(loan, "status", "active") == "closed":
            return None
        today = date.today()
        from_date = today
        to_date = today + timedelta(days=365 * 2)
        due_dates = _get_due_dates_range(loan, from_date, to_date)
        paid = await self._get_paid_due_dates_for_loan(loan_id)
        for due_dt in due_dates:
            if due_dt.date() not in paid:
                return (due_dt, loan.bi_weekly_payment_amount)
        return None

    async def validate_due_date_for_loan(
        self, loan_id: UUID, customer_id: UUID, due_dt: datetime, only_completed: bool = False
    ) -> tuple[datetime, float] | None:
        """Check that due_dt is a valid unpaid due date for the loan. Returns (due_dt, amount) or None.
        If only_completed=True, only completed payments count as paid (allows manual record for failed-due dates)."""
        loan = await self.db.get(Loan, loan_id)
        if not loan or loan.customer_id != customer_id:
            return None
        if getattr(loan, "status", "active") == "closed":
            return None
        due_d = due_dt.date()
        paid = await (
            self._get_paid_due_dates_for_loan_completed(loan_id)
            if only_completed
            else self._get_paid_due_dates_for_loan(loan_id)
        )
        if due_d in paid:
            return None
        # Check due_d is one of the scheduled due dates (window works for bi_weekly, monthly, semi_monthly)
        from_date = due_d - timedelta(days=60)
        to_date = due_d + timedelta(days=60)
        scheduled = _get_due_dates_range(loan, from_date, to_date)
        for s in scheduled:
            if s.date() == due_d:
                return (s, loan.bi_weekly_payment_amount)
        return None

    async def make_payment(
        self,
        customer_id: UUID,
        loan_id: UUID,
        card_token: str,
        payment_amount: float,
    ) -> dict:
        """
        Flexible payment: pay any amount. Applied to earliest dues first.
        Returns dict with success, message, transaction, charged_amount, requested_amount, excess_ignored.
        """
        if not settings.STRIPE_SECRET_KEY or not settings.STRIPE_SECRET_KEY.strip():
            AppException().raise_400("Stripe is not configured")
        if payment_amount <= 0:
            AppException().raise_400("payment_amount must be greater than 0")

        # Lock loan row for concurrency
        loan_row = await self.db.execute(
            select(Loan).where(Loan.id == loan_id).with_for_update()
        )
        loan = loan_row.scalar_one_or_none()
        if not loan:
            AppException().raise_404("Loan not found")
        if loan.customer_id != customer_id:
            AppException().raise_403("You can only pay for your own loan")
        if getattr(loan, "status", "active") == "closed":
            AppException().raise_400("Loan already completed")

        customer = await self.db.get(Customer, customer_id)
        if not customer:
            AppException().raise_404("Customer not found")

        remaining_balance = _round_amount(float(getattr(loan, "amount_financed", 0)))
        if remaining_balance <= 0:
            AppException().raise_400("Loan already completed")

        charge_amount = _round_amount(min(payment_amount, remaining_balance))
        amount_cents = int(round(charge_amount * 100))
        if amount_cents < 50:
            AppException().raise_400("Amount too small for Stripe (minimum $0.50)")

        try:
            result = create_payment_intent(
                amount_cents=amount_cents,
                currency=settings.STRIPE_CURRENCY,
                loan_id=loan_id,
                customer_id=customer_id,
                due_date_iso="flexible",
                customer_email=customer.email,
            )
            pi = confirm_payment_intent_with_token(result["payment_intent_id"], card_token)
        except Exception as e:
            err_msg = str(e).strip().lower()
            if "no such paymentmethod" in err_msg or "_secret_" in err_msg:
                AppException().raise_400(
                    "Invalid card_token: send a PaymentMethod ID (pm_xxx) or token (tok_xxx). "
                    "Do not send the PaymentIntent client_secret."
                )
            AppException().raise_400(f"Payment failed: {e}")

        if pi.status != "succeeded":
            AppException().raise_400(f"Payment not completed (status: {pi.status})")

        amount_received = _round_amount(
            (pi.get("amount_received") or pi.get("amount") or 0) / 100.0
        )
        paid_per = await self._get_paid_amount_per_due_date(loan_id, completed_only=True)
        today = date.today()
        loan_start = (loan.created_at + timedelta(days=14)).date()
        loan_end = loan.created_at + timedelta(days=int(loan.loan_term_months * 30.44))
        to_date = loan_end.date() if loan_end else today + timedelta(days=365 * 2)
        due_dates = _get_due_dates_range(loan, loan_start, to_date)
        amt_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
        installments: list[tuple[datetime, float, float, float]] = []
        for due_dt in due_dates:
            d = due_dt.date()
            paid = paid_per.get(d, 0)
            remaining = _round_amount(max(0, amt_per - paid))
            if remaining > 0:
                installments.append((due_dt, amt_per, paid, remaining))

        remaining_to_apply = amount_received
        applied_breakdown: list[dict] = []
        first_due = None
        for due_dt, _amt, _paid, rem in installments:
            if remaining_to_apply <= 0:
                break
            applied = _round_amount(min(remaining_to_apply, rem))
            if applied > 0:
                applied_breakdown.append({
                    "due_date": due_dt.strftime("%Y-%m-%dT%H:%M:%S"),
                    "applied_amount": applied,
                })
                if first_due is None:
                    first_due = due_dt
                remaining_to_apply -= applied

        loan.amount_financed = _round_amount(max(0, float(loan.amount_financed) - amount_received))
        loan.total_paid = _round_amount(float(getattr(loan, "total_paid", 0)) + amount_received)
        if loan.amount_financed <= 0:
            loan.amount_financed = 0
            loan.status = "closed"
        self.db.add(loan)

        payment = await self._record_flexible_payment(
            loan_id=loan_id,
            customer_id=customer_id,
            amount=amount_received,
            applied_installments=applied_breakdown,
            first_due=first_due or (due_dates[0] if due_dates else datetime.utcnow()),
            status="completed",
        )
        await self.db.commit()

        transaction = None
        if payment:
            await self.db.refresh(payment)
            vehicle = await self.db.get(Vehicle, loan.vehicle_id)
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
            customer_name = f"{customer.first_name} {customer.last_name}"
            transaction = TransactionItem(
                id=payment.id,
                loan_id=payment.loan_id,
                customer_id=payment.customer_id,
                customer_name=customer_name,
                vehicle_display=vehicle_display,
                amount=_ensure_non_negative_amount(payment.amount),
                emi_amount=_ensure_non_negative_amount(loan.bi_weekly_payment_amount),
                payment_method=payment.payment_method,
                status=payment.status,
                payment_date=payment.payment_date,
                due_date=payment.due_date,
                created_at=payment.created_at,
                loan_status=_loan_status_display(loan),
            )
            await send_payment_notification(
                self.db,
                customer_id=customer_id,
                notification_type="payment_received",
                scope_key=scope_key_for_payment(payment.id),
                title="Payment received",
                body=f"Your payment of ${_ensure_non_negative_amount(payment.amount):.2f} has been received.",
            )

        excess = _round_amount(max(0, payment_amount - charge_amount))
        return {
            "success": True,
            "message": "Payment completed successfully",
            "transaction": transaction,
            "charged_amount": charge_amount,
            "requested_amount": _round_amount(payment_amount),
            "excess_ignored": excess if excess > 0 else None,
        }

    async def get_checkout(
        self,
        loan_id: UUID,
        payment_amount: float,
        email: str | None,
        customer_id: UUID | None,
    ) -> dict:
        """
        Flexible checkout: pay any amount. Resolve customer, validate loan, create Stripe PaymentIntent.
        charge_amount = min(payment_amount, remaining_balance). No auth required.
        """
        if not settings.STRIPE_SECRET_KEY or not settings.STRIPE_SECRET_KEY.strip():
            AppException().raise_400("Stripe is not configured")
        if payment_amount <= 0:
            AppException().raise_400("payment_amount must be greater than 0")
        customer = None
        if customer_id:
            customer = await self.db.get(Customer, customer_id)
        if not customer and email and email.strip():
            result = await self.db.execute(select(Customer).where(Customer.email == email.strip()))
            customer = result.scalars().first()
        if not customer:
            AppException().raise_404("Customer not found. Provide a valid email or customer_id.")
        customer_id_resolved = customer.id
        loan = await self.db.get(Loan, loan_id)
        if not loan:
            AppException().raise_404("Loan not found")
        if loan.customer_id != customer_id_resolved:
            AppException().raise_403("This loan does not belong to the given customer")
        if getattr(loan, "status", "active") == "closed":
            AppException().raise_400("Loan already completed")
        remaining_balance = _round_amount(float(getattr(loan, "amount_financed", 0)))
        if remaining_balance <= 0:
            AppException().raise_400("Loan already completed")
        charge_amount = _round_amount(min(payment_amount, remaining_balance))
        amount_cents = int(round(charge_amount * 100))
        if amount_cents < 50:
            AppException().raise_400("Amount too small for Stripe (minimum $0.50)")
        try:
            result = create_payment_intent(
                amount_cents=amount_cents,
                currency=settings.STRIPE_CURRENCY,
                loan_id=loan_id,
                customer_id=customer_id_resolved,
                due_date_iso="flexible",
                customer_email=customer.email,
            )
        except Exception as e:
            AppException().raise_400(f"Checkout failed: {e}")
        vehicle = await self.db.get(Vehicle, loan.vehicle_id)
        vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
        customer_name = f"{customer.first_name} {customer.last_name}"
        first_due = None
        today = date.today()
        loan_start = (loan.created_at + timedelta(days=14)).date()
        loan_end = loan.created_at + timedelta(days=int(loan.loan_term_months * 30.44))
        to_date = loan_end.date() if loan_end else today + timedelta(days=365 * 2)
        due_dates = _get_due_dates_range(loan, loan_start, to_date)
        if due_dates:
            first_due = due_dates[0]
        return {
            "amount": charge_amount,
            "emi_amount": charge_amount,
            "amount_cents": amount_cents,
            "currency": settings.STRIPE_CURRENCY or "usd",
            "client_secret": result["client_secret"],
            "payment_intent_id": result["payment_intent_id"],
            "due_date": first_due or datetime.utcnow(),
            "customer_name": customer_name,
            "vehicle_display": vehicle_display,
            "loan_id": loan_id,
            "customer_id": customer_id_resolved,
            "loan_status": _loan_status_display(loan),
        }

    async def get_checkout_by_payment_intent_id(self, payment_intent_id: str) -> dict:
        """
        Retrieve checkout details by Stripe PaymentIntent ID (e.g. for polling or status).
        Returns amount, status, client_secret, metadata-derived fields, and optional display info from DB.
        """
        if not settings.STRIPE_SECRET_KEY or not settings.STRIPE_SECRET_KEY.strip():
            AppException().raise_400("Stripe is not configured")
        intent = get_payment_intent(payment_intent_id)
        if not intent:
            AppException().raise_404("Checkout not found. Invalid or expired payment_intent_id.")
        amount_cents = int(intent.get("amount") or 0)
        amount = amount_cents / 100.0
        currency = (intent.get("currency") or "usd").lower()
        status = intent.get("status") or "unknown"
        client_secret = intent.get("client_secret") or ""
        meta = intent.get("metadata") or {}
        loan_id = None
        customer_id = None
        due_date = None
        if meta.get("loan_id"):
            try:
                loan_id = UUID(meta["loan_id"])
            except (ValueError, TypeError):
                pass
        if meta.get("customer_id"):
            try:
                customer_id = UUID(meta["customer_id"])
            except (ValueError, TypeError):
                pass
        if meta.get("due_date_iso"):
            try:
                due_date = _parse_due_date_iso(meta["due_date_iso"])
            except (ValueError, TypeError):
                pass
        customer_name = None
        vehicle_display = None
        loan = None
        if customer_id and loan_id:
            customer = await self.db.get(Customer, customer_id)
            if customer:
                customer_name = f"{customer.first_name} {customer.last_name}"
            loan = await self.db.get(Loan, loan_id)
            if loan:
                vehicle = await self.db.get(Vehicle, loan.vehicle_id) if loan else None
                vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
        return {
            "amount": amount,
            "emi_amount": amount,
            "amount_cents": amount_cents,
            "currency": currency,
            "status": status,
            "client_secret": client_secret or None,
            "payment_intent_id": payment_intent_id,
            "due_date": due_date,
            "customer_name": customer_name,
            "vehicle_display": vehicle_display,
            "loan_id": loan_id,
            "customer_id": customer_id,
            "loan_status": _loan_status_display(loan),
        }

    async def confirm_public_payment(self, payment_intent_id: str, card_token: str) -> dict:
        """
        Confirm a PaymentIntent (from checkout) with card_token, record payment, return transaction.
        Handles flexible checkout (due_date_iso=flexible): applies to installments, updates loan.
        No auth required.
        """
        if not settings.STRIPE_SECRET_KEY or not settings.STRIPE_SECRET_KEY.strip():
            AppException().raise_400("Stripe is not configured")
        try:
            pi = confirm_payment_intent_with_token(payment_intent_id, card_token)
        except Exception as e:
            err_msg = str(e).strip().lower()
            if "no such paymentmethod" in err_msg or "_secret_" in err_msg:
                AppException().raise_400(
                    "Invalid card_token: send a PaymentMethod ID (pm_xxx) or token (tok_xxx). "
                    "Do not send the PaymentIntent client_secret."
                )
            AppException().raise_400(f"Payment failed: {e}")
        if pi.status != "succeeded":
            AppException().raise_400(f"Payment not completed (status: {pi.status})")
        meta = pi.get("metadata") or {}
        loan_id_str = meta.get("loan_id")
        customer_id_str = meta.get("customer_id")
        due_date_iso = meta.get("due_date_iso") or ""
        if not loan_id_str or not customer_id_str:
            AppException().raise_400("Invalid PaymentIntent: missing loan/customer metadata")
        loan_id = UUID(loan_id_str)
        customer_id = UUID(customer_id_str)
        amount_received = _round_amount(
            (pi.get("amount_received") or pi.get("amount") or 0) / 100.0
        )
        loan = await self.db.get(Loan, loan_id)
        if not loan:
            AppException().raise_400("Loan not found")
        if loan.customer_id != customer_id:
            AppException().raise_403("Loan does not belong to this customer")
        is_flexible = due_date_iso.strip().lower() == "flexible"
        if is_flexible:
            paid_per = await self._get_paid_amount_per_due_date(loan_id, completed_only=True)
            today = date.today()
            loan_start = (loan.created_at + timedelta(days=14)).date()
            loan_end = loan.created_at + timedelta(days=int(loan.loan_term_months * 30.44))
            to_date = loan_end.date() if loan_end else today + timedelta(days=365 * 2)
            due_dates = _get_due_dates_range(loan, loan_start, to_date)
            amt_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            installments: list[tuple[datetime, float, float, float]] = []
            for due_dt in due_dates:
                d = due_dt.date()
                paid = paid_per.get(d, 0)
                remaining = _round_amount(max(0, amt_per - paid))
                if remaining > 0:
                    installments.append((due_dt, amt_per, paid, remaining))
            remaining_to_apply = amount_received
            applied_breakdown: list[dict] = []
            first_due = None
            for due_dt, _amt, _paid, rem in installments:
                if remaining_to_apply <= 0:
                    break
                applied = _round_amount(min(remaining_to_apply, rem))
                if applied > 0:
                    applied_breakdown.append({
                        "due_date": due_dt.strftime("%Y-%m-%dT%H:%M:%S"),
                        "applied_amount": applied,
                    })
                    if first_due is None:
                        first_due = due_dt
                    remaining_to_apply -= applied
            loan.amount_financed = _round_amount(max(0, float(loan.amount_financed) - amount_received))
            loan.total_paid = _round_amount(float(getattr(loan, "total_paid", 0)) + amount_received)
            if loan.amount_financed <= 0:
                loan.amount_financed = 0
                loan.status = "closed"
            self.db.add(loan)
            payment = await self._record_flexible_payment(
                loan_id=loan_id,
                customer_id=customer_id,
                amount=amount_received,
                applied_installments=applied_breakdown,
                first_due=first_due or (due_dates[0] if due_dates else datetime.utcnow()),
                status="completed",
                payment_mode=PaymentMode.checkout.value,
            )
            await self.db.commit()
            await self.db.refresh(payment)
        else:
            try:
                due_dt = _parse_due_date_iso(due_date_iso)
            except (ValueError, TypeError):
                AppException().raise_400("Invalid due_date in PaymentIntent metadata")
            emi_amt = _ensure_non_negative_amount(loan.bi_weekly_payment_amount) if loan else amount_received
            payment = await self._record_payment(
                loan_id=loan_id,
                customer_id=customer_id,
                due_date=due_dt,
                amount=amount_received,
                emi_amount=emi_amt,
                status="completed",
            )
            if not payment:
                existing = await self.db.execute(
                    select(Payment).where(
                        Payment.loan_id == loan_id,
                        func.date(Payment.due_date) == due_dt.date(),
                    )
                )
                payment = existing.scalars().first()
        transaction = None
        if payment:
            loan = await self.db.get(Loan, payment.loan_id)
            customer = await self.db.get(Customer, payment.customer_id)
            vehicle = await self.db.get(Vehicle, loan.vehicle_id) if loan else None
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
            customer_name = f"{customer.first_name} {customer.last_name}" if customer else None
            emi_amt_val = getattr(payment, "emi_amount", None) or (loan and _ensure_non_negative_amount(loan.bi_weekly_payment_amount)) or _ensure_non_negative_amount(payment.amount)
            transaction = TransactionItem(
                id=payment.id,
                loan_id=payment.loan_id,
                customer_id=payment.customer_id,
                customer_name=customer_name,
                vehicle_display=vehicle_display,
                amount=_ensure_non_negative_amount(payment.amount),
                emi_amount=_ensure_non_negative_amount(emi_amt_val),
                payment_method=payment.payment_method,
                status=payment.status,
                payment_date=payment.payment_date,
                due_date=payment.due_date,
                created_at=payment.created_at,
                loan_status=_loan_status_display(loan),
            )
            await send_payment_notification(
                self.db,
                customer_id=payment.customer_id,
                notification_type="payment_received",
                scope_key=scope_key_for_payment(payment.id),
                title="Payment received",
                body=f"Your payment of ${_ensure_non_negative_amount(payment.amount):.2f} has been received.",
            )
            await self._check_and_close_loan_if_paid(loan_id)
        return {
            "success": True,
            "message": "Payment completed successfully",
            "transaction": transaction,
        }

    async def _record_payment(
        self,
        loan_id: UUID,
        customer_id: UUID,
        due_date: datetime,
        amount: float,
        emi_amount: float | None = None,
        status: str = "completed",
    ) -> Payment | None:
        """Create payment record. Idempotent: returns None if already exists for this loan + due date."""
        existing = await self.db.execute(
            select(Payment).where(
                Payment.loan_id == loan_id,
                func.date(Payment.due_date) == due_date.date(),
            )
        )
        if existing.scalars().first():
            return None
        payment = Payment(
            loan_id=loan_id,
            customer_id=customer_id,
            amount=amount,
            emi_amount=emi_amount,
            payment_method="card",
            payment_date=datetime.utcnow(),
            due_date=due_date,
            status=status,
            payment_mode=PaymentMode.installment.value,
        )
        self.db.add(payment)
        await self.db.commit()
        await self.db.refresh(payment)
        return payment

    async def _record_flexible_payment(
        self,
        loan_id: UUID,
        customer_id: UUID,
        amount: float,
        applied_installments: list[dict],
        first_due: datetime,
        status: str = "completed",
        payment_mode: str = PaymentMode.manual.value,
    ) -> Payment | None:
        """Create flexible payment record with applied_installments breakdown."""
        payment = Payment(
            loan_id=loan_id,
            customer_id=customer_id,
            amount=amount,
            emi_amount=None,
            payment_method="card",
            payment_date=datetime.utcnow(),
            due_date=first_due,
            status=status,
            payment_mode=payment_mode,
            applied_installments=applied_installments if applied_installments else None,
        )
        self.db.add(payment)
        return payment

    async def _record_manual_payment(
        self,
        loan_id: UUID,
        customer_id: UUID,
        due_date: datetime,
        amount: float,
        payment_method: str,
        emi_amount: float | None = None,
        note: str | None = None,
    ) -> Payment | None:
        """Create manual payment record. Returns None if duplicate (completed payment already exists for loan + due date)."""
        existing = await self.db.execute(
            select(Payment).where(
                Payment.loan_id == loan_id,
                func.date(Payment.due_date) == due_date.date(),
                Payment.status == PaymentStatus.completed.value,
            )
        )
        if existing.scalars().first():
            return None
        payment = Payment(
            loan_id=loan_id,
            customer_id=customer_id,
            amount=amount,
            emi_amount=emi_amount,
            payment_method=payment_method,
            payment_date=datetime.utcnow(),
            due_date=due_date,
            status=PaymentStatus.completed.value,
            payment_mode=PaymentMode.installment.value,
            note=note,
        )
        self.db.add(payment)
        await self.db.commit()
        await self.db.refresh(payment)
        return payment

    async def _check_and_close_loan_if_paid(self, loan_id: UUID) -> None:
        """If amount_financed is zero or total completed payments >= amount_financed, set loan status to closed."""
        loan = await self.db.get(Loan, loan_id)
        if not loan or getattr(loan, "status", "active") == "closed":
            return
        amount_financed = float(loan.amount_financed)
        if amount_financed <= 0:
            loan.status = "closed"
            self.db.add(loan)
            await self.db.commit()
            return
        total_result = await self.db.execute(
            select(func.coalesce(func.sum(Payment.amount), 0)).where(
                Payment.loan_id == loan_id,
                Payment.status == PaymentStatus.completed.value,
            )
        )
        total_paid = float(total_result.scalar_one() or 0)
        if total_paid >= amount_financed - 0.01:
            loan.status = "closed"
            self.db.add(loan)
            await self.db.commit()

    async def waive_overdue_installment_admin(
        self,
        loan_id: UUID,
        due_date_iso: str,
        note: str | None = None,
    ) -> TransactionItem | None:
        """Waive an unpaid installment: create a zero-amount completed payment with payment_method=waived. Returns TransactionItem or None if invalid."""
        loan = await self.db.get(Loan, loan_id)
        if not loan:
            return None
        customer_id = loan.customer_id
        try:
            due_dt = _parse_due_date_iso(due_date_iso)
        except ValueError:
            return None
        validated = await self.validate_due_date_for_loan(loan_id, customer_id, due_dt, only_completed=True)
        if not validated:
            return None  # Invalid or already paid/waived due date
        emi_amt = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
        payment = await self._record_manual_payment(
            loan_id=loan_id,
            customer_id=customer_id,
            due_date=due_dt,
            amount=0.0,
            payment_method="waived",
            emi_amount=emi_amt,
            note=note or "Installment waived by admin",
        )
        if not payment:
            return None
        await self._check_and_close_loan_if_paid(loan_id)
        customer = await self.db.get(Customer, customer_id)
        vehicle = await self.db.get(Vehicle, loan.vehicle_id)
        customer_name = f"{customer.first_name} {customer.last_name}" if customer else None
        vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
        return TransactionItem(
            id=payment.id,
            loan_id=payment.loan_id,
            customer_id=payment.customer_id,
            customer_name=customer_name,
            vehicle_display=vehicle_display,
            amount=0.0,
            emi_amount=emi_amt,
            payment_method=payment.payment_method,
            status=payment.status,
            payment_date=payment.payment_date,
            due_date=payment.due_date,
            created_at=payment.created_at,
            loan_status=_loan_status_display(loan),
        )

    async def waive_earliest_overdue_by_customer_loan(
        self,
        customer_id: UUID,
        loan_id: UUID,
        note: str | None = None,
    ) -> TransactionItem | None:
        """Waive the earliest overdue installment for the given customer and loan. Returns TransactionItem or None if no overdue or invalid."""
        loan = await self.db.get(Loan, loan_id)
        if not loan or loan.customer_id != customer_id:
            return None
        if getattr(loan, "status", "active") == "closed":
            return None
        today = date.today()
        from_date = (loan.created_at + timedelta(days=14)).date()
        to_date = today
        due_dates = _get_due_dates_range(loan, from_date, to_date)
        paid_completed = await self._get_paid_due_dates_for_loan_completed(loan.id)
        earliest_overdue_dt = None
        for due_dt in due_dates:
            due_d = due_dt.date()
            if due_d >= today or due_d in paid_completed:
                continue
            earliest_overdue_dt = due_dt
            break
        if earliest_overdue_dt is None:
            return None
        due_date_iso = earliest_overdue_dt.isoformat()
        return await self.waive_overdue_installment_admin(
            loan_id=loan_id,
            due_date_iso=due_date_iso,
            note=note,
        )

    async def record_manual_payment_admin(
        self,
        customer_id: UUID,
        loan_id: UUID,
        due_date_iso: str,
        amount: float,
        payment_method: str,
        note: str | None = None,
    ) -> TransactionItem | None:
        """Admin records a manual payment received from a customer. Returns TransactionItem or None if invalid."""
        loan = await self.db.get(Loan, loan_id)
        if not loan:
            return None
        if loan.customer_id != customer_id:
            return None
        try:
            due_dt = _parse_due_date_iso(due_date_iso)
        except ValueError:
            return None
        validated = await self.validate_due_date_for_loan(loan_id, customer_id, due_dt, only_completed=True)
        if not validated:
            return None  # Invalid or already paid due date
        amount_safe = _ensure_non_negative_amount(amount)
        emi_amt = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
        payment = await self._record_manual_payment(
            loan_id=loan_id,
            customer_id=customer_id,
            due_date=due_dt,
            amount=amount_safe,
            payment_method=payment_method,
            emi_amount=emi_amt,
            note=note,
        )
        if not payment:
            return None
        loan.amount_financed = _round_amount(max(0, float(loan.amount_financed) - amount_safe))
        loan.total_paid = _round_amount(float(getattr(loan, "total_paid", 0)) + amount_safe)
        if loan.amount_financed <= 0:
            loan.amount_financed = 0
            loan.status = "closed"
        self.db.add(loan)
        await self.db.commit()
        customer = await self.db.get(Customer, customer_id)
        vehicle = await self.db.get(Vehicle, loan.vehicle_id)
        customer_name = f"{customer.first_name} {customer.last_name}" if customer else None
        vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
        await send_payment_notification(
            self.db,
            customer_id=customer_id,
            notification_type="payment_received",
            scope_key=scope_key_for_payment(payment.id),
            title="Payment received",
            body=f"Your payment of ${_ensure_non_negative_amount(payment.amount):.2f} has been received.",
        )
        return TransactionItem(
            id=payment.id,
            loan_id=payment.loan_id,
            customer_id=payment.customer_id,
            customer_name=customer_name,
            vehicle_display=vehicle_display,
            amount=_ensure_non_negative_amount(payment.amount),
            emi_amount=emi_amt,
            payment_method=payment.payment_method,
            status=payment.status,
            payment_date=payment.payment_date,
            due_date=payment.due_date,
            created_at=payment.created_at,
            loan_status=_loan_status_display(loan),
        )

    async def list_my_transactions(
        self,
        customer_id: UUID,
        skip: int = 0,
        limit: int = 100,
    ) -> tuple[list[TransactionItem], int]:
        """Transaction history for the logged-in user. Returns (items, total)."""
        base = (
            select(Payment, Customer, Vehicle, Loan)
            .join(Loan, Payment.loan_id == Loan.id)
            .join(Customer, Payment.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
            .where(Payment.customer_id == customer_id)
            .order_by(Payment.payment_date.desc())
        )
        total_result = await self.db.execute(
            select(func.count(Payment.id)).where(Payment.customer_id == customer_id)
        )
        total = total_result.scalar_one() or 0
        result = await self.db.execute(base.offset(skip).limit(limit))
        items = []
        for p, c, v, loan in result.all():
            emi_val = getattr(p, "emi_amount", None)
            if emi_val is None:
                emi_val = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            items.append(
                TransactionItem(
                    id=p.id,
                    loan_id=p.loan_id,
                    customer_id=p.customer_id,
                    customer_name=f"{c.first_name} {c.last_name}",
                    vehicle_display=f"{v.year} {v.make} {v.model}" if v else None,
                    amount=_ensure_non_negative_amount(p.amount),
                    emi_amount=_ensure_non_negative_amount(emi_val),
                    payment_method=p.payment_method,
                    status=p.status,
                    payment_date=p.payment_date,
                    due_date=p.due_date,
                    created_at=p.created_at,
                    loan_status=_loan_status_display(loan),
                )
            )
        return items, total

    async def list_transactions_admin(
        self,
        customer_id: UUID | None = None,
        loan_id: UUID | None = None,
        from_date: date | None = None,
        to_date: date | None = None,
        skip: int = 0,
        limit: int = 100,
    ) -> tuple[list[TransactionItem], int, float, int, int]:
        """Transaction history for all users with filters. Returns (items, total, total_amount, completed_count, failed_count)."""
        base = (
            select(Payment, Customer, Vehicle, Loan)
            .join(Loan, Payment.loan_id == Loan.id)
            .join(Customer, Payment.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
            .order_by(Payment.payment_date.desc())
        )
        count_q = select(func.count(Payment.id))
        sum_q = select(
            func.coalesce(
                func.sum(case((Payment.amount < 0, 0), else_=Payment.amount)), 0
            )
        )
        completed_q = select(func.count(Payment.id)).where(Payment.status == PaymentStatus.completed.value)
        failed_q = select(func.count(Payment.id)).where(Payment.status == "failed")
        if customer_id is not None:
            base = base.where(Payment.customer_id == customer_id)
            count_q = count_q.where(Payment.customer_id == customer_id)
            sum_q = sum_q.where(Payment.customer_id == customer_id)
            completed_q = completed_q.where(Payment.customer_id == customer_id)
            failed_q = failed_q.where(Payment.customer_id == customer_id)
        if loan_id is not None:
            base = base.where(Payment.loan_id == loan_id)
            count_q = count_q.where(Payment.loan_id == loan_id)
            sum_q = sum_q.where(Payment.loan_id == loan_id)
            completed_q = completed_q.where(Payment.loan_id == loan_id)
            failed_q = failed_q.where(Payment.loan_id == loan_id)
        if from_date is not None:
            base = base.where(func.date(Payment.payment_date) >= from_date)
            count_q = count_q.where(func.date(Payment.payment_date) >= from_date)
            sum_q = sum_q.where(func.date(Payment.payment_date) >= from_date)
            completed_q = completed_q.where(func.date(Payment.payment_date) >= from_date)
            failed_q = failed_q.where(func.date(Payment.payment_date) >= from_date)
        if to_date is not None:
            base = base.where(func.date(Payment.payment_date) <= to_date)
            count_q = count_q.where(func.date(Payment.payment_date) <= to_date)
            sum_q = sum_q.where(func.date(Payment.payment_date) <= to_date)
            completed_q = completed_q.where(func.date(Payment.payment_date) <= to_date)
            failed_q = failed_q.where(func.date(Payment.payment_date) <= to_date)

        total_result = await self.db.execute(count_q)
        total = total_result.scalar_one() or 0
        sum_result = await self.db.execute(sum_q)
        total_amount = float(sum_result.scalar_one() or 0)
        completed_result = await self.db.execute(completed_q)
        completed_count = completed_result.scalar_one() or 0
        failed_result = await self.db.execute(failed_q)
        failed_count = failed_result.scalar_one() or 0

        result = await self.db.execute(base.offset(skip).limit(limit))
        items = []
        for p, c, v, loan in result.all():
            emi_val = getattr(p, "emi_amount", None)
            if emi_val is None:
                emi_val = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            items.append(
                TransactionItem(
                    id=p.id,
                    loan_id=p.loan_id,
                    customer_id=p.customer_id,
                    customer_name=f"{c.first_name} {c.last_name}",
                    vehicle_display=f"{v.year} {v.make} {v.model}" if v else None,
                    amount=_ensure_non_negative_amount(p.amount),
                    emi_amount=_ensure_non_negative_amount(emi_val),
                    payment_method=p.payment_method,
                    status=p.status,
                    payment_date=p.payment_date,
                    due_date=p.due_date,
                    created_at=p.created_at,
                    loan_status=_loan_status_display(loan),
                )
            )
        return items, total, total_amount, completed_count, failed_count

    async def export_transactions_to_excel(
        self,
        customer_id: UUID | None = None,
        loan_id: UUID | None = None,
        from_date: date | None = None,
        to_date: date | None = None,
    ) -> bytes:
        """Export transactions to Excel (.xlsx). Same filters as list_transactions_admin (no pagination)."""
        base = (
            select(Payment, Customer, Vehicle, Loan)
            .join(Loan, Payment.loan_id == Loan.id)
            .join(Customer, Payment.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
            .order_by(Payment.payment_date.desc())
        )
        if customer_id is not None:
            base = base.where(Payment.customer_id == customer_id)
        if loan_id is not None:
            base = base.where(Payment.loan_id == loan_id)
        if from_date is not None:
            base = base.where(func.date(Payment.payment_date) >= from_date)
        if to_date is not None:
            base = base.where(func.date(Payment.payment_date) <= to_date)
        result = await self.db.execute(base)
        rows = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = [
            "Payment ID",
            "Loan ID",
            "Customer ID",
            "Customer Name",
            "Vehicle",
            "Amount",
            "EMI Amount",
            "Payment Method",
            "Status",
            "Payment Date",
            "Due Date",
            "Loan Status",
            "Created At",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, (p, c, v, loan) in enumerate(rows, start=2):
            customer_name = f"{c.first_name} {c.last_name}"
            vehicle_display = f"{v.year} {v.make} {v.model}" if v else ""
            emi_val = getattr(p, "emi_amount", None) or _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            payment_date_str = p.payment_date.strftime("%Y-%m-%d %H:%M:%S") if hasattr(p.payment_date, "strftime") else str(p.payment_date)
            due_date_str = p.due_date.strftime("%Y-%m-%d %H:%M:%S") if hasattr(p.due_date, "strftime") else str(p.due_date)
            created_str = p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at and hasattr(p.created_at, "strftime") else str(p.created_at or "")
            ws.cell(row=row_idx, column=1, value=str(p.id))
            ws.cell(row=row_idx, column=2, value=str(p.loan_id))
            ws.cell(row=row_idx, column=3, value=str(p.customer_id))
            ws.cell(row=row_idx, column=4, value=customer_name)
            ws.cell(row=row_idx, column=5, value=vehicle_display)
            ws.cell(row=row_idx, column=6, value=_ensure_non_negative_amount(p.amount))
            ws.cell(row=row_idx, column=7, value=_ensure_non_negative_amount(emi_val))
            ws.cell(row=row_idx, column=8, value=p.payment_method or "")
            ws.cell(row=row_idx, column=9, value=p.status or "")
            ws.cell(row=row_idx, column=10, value=payment_date_str)
            ws.cell(row=row_idx, column=11, value=due_date_str)
            ws.cell(row=row_idx, column=12, value=_loan_status_display(loan))
            ws.cell(row=row_idx, column=13, value=created_str)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def get_payment_summary_admin(
        self,
        customer_id: UUID | None = None,
        loan_id: UUID | None = None,
        search: str | None = None,
    ) -> dict:
        """
        Get payment summary: paid dues, unpaid dues, overdue payments, totals.
        search: filter by customer name, email, or phone (case-insensitive substring).
        """
        today = date.today()
        base_loans = (
            select(Loan, Customer, Vehicle)
            .join(Customer, Loan.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
        )
        if customer_id is not None:
            base_loans = base_loans.where(Loan.customer_id == customer_id)
        if loan_id is not None:
            base_loans = base_loans.where(Loan.id == loan_id)
        if search and search.strip():
            q = search.strip()
            base_loans = base_loans.where(
                or_(
                    Customer.first_name.ilike(f"%{q}%"),
                    Customer.last_name.ilike(f"%{q}%"),
                    Customer.email.ilike(f"%{q}%"),
                    Customer.phone.ilike(f"%{q}%"),
                    func.concat(Customer.first_name, " ", Customer.last_name).ilike(f"%{q}%"),
                )
            )
        loans_result = await self.db.execute(base_loans)
        paid_dues: list[DueEntryItem] = []
        unpaid_dues: list[DueEntryItem] = []
        overdue_payments: list[DueEntryItem] = []
        total_collected = 0.0
        pending_amount = 0.0
        overdue_amount = 0.0

        for loan, customer, vehicle in loans_result.all():
            customer_name = f"{customer.first_name} {customer.last_name}" if customer else None
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
            loan_status = _loan_status_display(loan)
            is_closed = getattr(loan, "status", "active") == "closed"

            # Paid dues: from Payment records (including flexible with applied_installments)
            paid_result = await self.db.execute(
                select(Payment)
                .where(
                    Payment.loan_id == loan.id,
                    Payment.status == PaymentStatus.completed.value,
                )
                .order_by(Payment.due_date.asc())
            )
            emi_amt = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            for p in paid_result.scalars().all():
                amt = _ensure_non_negative_amount(p.amount)
                total_collected += amt
                emi_val = getattr(p, "emi_amount", None) or emi_amt
                applied = getattr(p, "applied_installments", None) if isinstance(getattr(p, "applied_installments", None), list) else None
                if applied:
                    for entry in applied:
                        if isinstance(entry, dict):
                            due_str = entry.get("due_date")
                            app_amt = float(entry.get("applied_amount") or 0)
                            if due_str and app_amt > 0:
                                try:
                                    due_dt = datetime.fromisoformat(due_str.replace("Z", "+00:00"))
                                    paid_dues.append(
                                        DueEntryItem(
                                            loan_id=loan.id,
                                            customer_id=loan.customer_id,
                                            customer_name=customer_name,
                                            vehicle_display=vehicle_display,
                                            due_date=due_dt,
                                            amount=app_amt,
                                            emi_amount=emi_amt,
                                            payment_date=p.payment_date,
                                            payment_id=p.id,
                                            days_overdue=None,
                                            days_until_due=None,
                                            loan_status=loan_status,
                                        )
                                    )
                                except (ValueError, TypeError):
                                    pass
                else:
                    paid_dues.append(
                        DueEntryItem(
                            loan_id=loan.id,
                            customer_id=loan.customer_id,
                            customer_name=customer_name,
                            vehicle_display=vehicle_display,
                            due_date=p.due_date,
                            amount=amt,
                            emi_amount=_ensure_non_negative_amount(emi_val),
                            payment_date=p.payment_date,
                            payment_id=p.id,
                            days_overdue=None,
                            days_until_due=None,
                            loan_status=loan_status,
                        )
                    )

            # For closed loans, do not add unpaid or overdue
            if is_closed:
                continue

            # Unpaid/overdue: use paid_amount per due_date (supports partial payments)
            loan_start = (loan.created_at + timedelta(days=14)).date()
            loan_end = loan.created_at + timedelta(days=int(loan.loan_term_months * 30.44))
            to_date = loan_end.date() if loan_end else today + timedelta(days=365 * 2)
            all_due_dates = _get_due_dates_range(loan, loan_start, to_date)
            paid_per = await self._get_paid_amount_per_due_date(loan.id, completed_only=True)
            amount_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)

            for due_dt in all_due_dates:
                due_d = due_dt.date()
                paid_amt = paid_per.get(due_d, 0)
                remaining = _round_amount(max(0, amount_per - paid_amt))
                if remaining <= 0:
                    continue
                if due_d < today:
                    days_overdue = (today - due_d).days
                    overdue_amount += remaining
                    overdue_payments.append(
                        DueEntryItem(
                            loan_id=loan.id,
                            customer_id=loan.customer_id,
                            customer_name=customer_name,
                            vehicle_display=vehicle_display,
                            due_date=due_dt,
                            amount=remaining,
                            emi_amount=amount_per,
                            payment_date=None,
                            payment_id=None,
                            days_overdue=days_overdue,
                            days_until_due=None,
                            loan_status=loan_status,
                        )
                    )
                else:
                    days_until_due = (due_d - today).days
                    pending_amount += remaining
                    unpaid_dues.append(
                        DueEntryItem(
                            loan_id=loan.id,
                            customer_id=loan.customer_id,
                            customer_name=customer_name,
                            vehicle_display=vehicle_display,
                            due_date=due_dt,
                            amount=remaining,
                            emi_amount=amount_per,
                            payment_date=None,
                            payment_id=None,
                            days_overdue=None,
                            days_until_due=days_until_due,
                            loan_status=loan_status,
                        )
                    )

        overdue_payments.sort(key=lambda x: (x.due_date, x.loan_id))
        unpaid_dues.sort(key=lambda x: (x.due_date, x.loan_id))
        paid_dues.sort(key=lambda x: (x.due_date, x.loan_id))

        # Single combined list with payment_status: paid_dues | unpaid_dues | overdue_payments
        items = []
        for p in paid_dues:
            items.append({"payment_status": "paid_dues", **p.model_dump()})
        for o in overdue_payments:
            items.append({"payment_status": "overdue_payments", **o.model_dump()})
        for u in unpaid_dues:
            items.append({"payment_status": "unpaid_dues", **u.model_dump()})
        items.sort(key=lambda x: (x["due_date"], str(x["loan_id"])))

        return {
            "items": items,
            "total_collected_amount": total_collected,
            "pending_amount": pending_amount,
            "overdue_amount": overdue_amount,
            "total_payment_left": pending_amount + overdue_amount,
        }

    async def list_due_customers_admin(
        self,
        customer_id: UUID | None = None,
        loan_id: UUID | None = None,
        search: str | None = None,
        skip: int = 0,
        limit: int = 500,
    ) -> tuple[list[DueCustomerItem], int]:
        """
        List customers who have at least one unpaid due, with loan_id and details for create checkout.
        One row per (customer, loan). Use loan_id + customer_id/email to call create checkout (payment_type=next or due).
        Returns (items_page, total).
        """
        today = date.today()
        base_loans = (
            select(Loan, Customer, Vehicle)
            .join(Customer, Loan.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
        )
        if customer_id is not None:
            base_loans = base_loans.where(Loan.customer_id == customer_id)
        if loan_id is not None:
            base_loans = base_loans.where(Loan.id == loan_id)
        if search and search.strip():
            q = search.strip()
            base_loans = base_loans.where(
                or_(
                    Customer.first_name.ilike(f"%{q}%"),
                    Customer.last_name.ilike(f"%{q}%"),
                    Customer.email.ilike(f"%{q}%"),
                    Customer.phone.ilike(f"%{q}%"),
                    func.concat(Customer.first_name, " ", Customer.last_name).ilike(f"%{q}%"),
                )
            )
        base_loans = base_loans.where(Loan.status != "closed")
        loans_result = await self.db.execute(base_loans)
        all_rows: list[DueCustomerItem] = []
        for loan, customer, vehicle in loans_result.all():
            loan_start = (loan.created_at + timedelta(days=14)).date()
            loan_end = loan.created_at + timedelta(days=int(loan.loan_term_months * 30.44))
            to_date = loan_end.date() if loan_end else today + timedelta(days=365 * 2)
            due_dates = _get_due_dates_range(loan, loan_start, to_date)
            paid_per = await self._get_paid_amount_per_due_date(loan.id, completed_only=True)
            unpaid_dates: list[datetime] = []
            total_unpaid = 0.0
            amount_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            for due_dt in due_dates:
                paid_amt = paid_per.get(due_dt.date(), 0)
                remaining = _round_amount(max(0, amount_per - paid_amt))
                if remaining <= 0:
                    continue
                unpaid_dates.append(due_dt)
                total_unpaid += remaining
            if not unpaid_dates:
                continue
            unpaid_dates.sort()
            next_due = unpaid_dates[0]
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
            all_rows.append(
                DueCustomerItem(
                    customer_id=customer.id,
                    email=customer.email,
                    customer_name=f"{customer.first_name} {customer.last_name}",
                    phone=customer.phone or None,
                    loan_id=loan.id,
                    vehicle_display=vehicle_display,
                    unpaid_count=len(unpaid_dates),
                    total_unpaid_amount=total_unpaid,
                    emi_amount=amount_per,
                    next_due_date=next_due,
                    next_due_date_iso=next_due.isoformat(),
                    loan_status=_loan_status_display(loan),
                )
            )
        total = len(all_rows)
        page = all_rows[skip : skip + limit]
        return (page, total)

    async def list_due_installments_admin(
        self,
        customer_id: UUID | None = None,
        loan_id: UUID | None = None,
        search: str | None = None,
        skip: int = 0,
        limit: int = 500,
    ) -> tuple[list[DueInstallmentItem], int, float]:
        """
        List every unpaid due installment with loan_id, customer, due_date_iso, amount for create checkout.
        Use each item to call create checkout with payment_type=due and due_date_iso.
        Returns (items_page, total, total_amount).
        """
        today = date.today()
        base_loans = (
            select(Loan, Customer, Vehicle)
            .join(Customer, Loan.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
        )
        if customer_id is not None:
            base_loans = base_loans.where(Loan.customer_id == customer_id)
        if loan_id is not None:
            base_loans = base_loans.where(Loan.id == loan_id)
        if search and search.strip():
            q = search.strip()
            base_loans = base_loans.where(
                or_(
                    Customer.first_name.ilike(f"%{q}%"),
                    Customer.last_name.ilike(f"%{q}%"),
                    Customer.email.ilike(f"%{q}%"),
                    Customer.phone.ilike(f"%{q}%"),
                    func.concat(Customer.first_name, " ", Customer.last_name).ilike(f"%{q}%"),
                )
            )
        base_loans = base_loans.where(Loan.status != "closed")
        loans_result = await self.db.execute(base_loans)
        all_items: list[DueInstallmentItem] = []
        total_amount = 0.0
        for loan, customer, vehicle in loans_result.all():
            loan_start = (loan.created_at + timedelta(days=14)).date()
            loan_end = loan.created_at + timedelta(days=int(loan.loan_term_months * 30.44))
            to_date = loan_end.date() if loan_end else today + timedelta(days=365 * 2)
            due_dates = _get_due_dates_range(loan, loan_start, to_date)
            paid_per = await self._get_paid_amount_per_due_date(loan.id, completed_only=True)
            amount_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
            customer_name = f"{customer.first_name} {customer.last_name}"
            loan_status = _loan_status_display(loan)
            for due_dt in due_dates:
                due_d = due_dt.date()
                paid_amt = paid_per.get(due_d, 0)
                remaining = _round_amount(max(0, amount_per - paid_amt))
                if remaining <= 0:
                    continue
                days_overdue = (today - due_d).days if due_d < today else None
                days_until_due = (due_d - today).days if due_d >= today else None
                all_items.append(
                    DueInstallmentItem(
                        loan_id=loan.id,
                        customer_id=customer.id,
                        email=customer.email,
                        customer_name=customer_name,
                        phone=customer.phone or None,
                        vehicle_display=vehicle_display,
                        due_date=due_dt,
                        due_date_iso=due_dt.isoformat(),
                        amount=remaining,
                        emi_amount=amount_per,
                        days_overdue=days_overdue,
                        days_until_due=days_until_due,
                        loan_status=loan_status,
                    )
                )
                total_amount += remaining
        all_items.sort(key=lambda x: (x.due_date, x.loan_id))
        total = len(all_items)
        page = all_items[skip : skip + limit]
        return (page, total, total_amount)

    async def list_overdue_for_admin(
        self,
        skip: int = 0,
        limit: int = 500,
    ) -> tuple[list[OverdueItem], int, float, float]:
        """
        List overdue installments (scheduled due dates in the past with no completed payment).
        Returns (items_page, total_overdue_count, total_outstanding_amount, avg_overdue_days).
        """
        today = date.today()
        loans_result = await self.db.execute(
            select(Loan, Customer, Vehicle)
            .join(Customer, Loan.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
            .where(Loan.status != "closed")
        )
        all_items: list[OverdueItem] = []
        for loan, customer, vehicle in loans_result.all():
            # Due dates from loan start up to (and including) yesterday
            from_date = loan.created_at.date()
            to_date = today
            due_dates = _get_due_dates_range(loan, from_date, to_date)
            paid_per = await self._get_paid_amount_per_due_date(loan.id, completed_only=True)
            amount_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            customer_name = f"{customer.first_name} {customer.last_name}" if customer else None
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
            for due_dt in due_dates:
                due_d = due_dt.date()
                if due_d >= today:
                    continue
                paid_amt = paid_per.get(due_d, 0)
                remaining = _round_amount(max(0, amount_per - paid_amt))
                if remaining <= 0:
                    continue
                days_overdue = (today - due_d).days
                all_items.append(
                    OverdueItem(
                        loan_id=loan.id,
                        customer_id=loan.customer_id,
                        customer_name=customer_name,
                        vehicle_display=vehicle_display,
                        due_date=due_dt,
                        amount=remaining,
                        emi_amount=amount_per,
                        days_overdue=days_overdue,
                        loan_status=_loan_status_display(loan),
                    )
                )
        # Sort by due_date ascending (oldest overdue first) or by days_overdue desc
        all_items.sort(key=lambda x: (x.due_date, x.loan_id))
        total_count = len(all_items)
        total_outstanding = sum(i.amount for i in all_items)
        avg_days = (sum(i.days_overdue for i in all_items) / total_count) if total_count else 0.0
        page = all_items[skip : skip + limit]
        return (page, total_count, total_outstanding, avg_days)

    async def get_overdue_customer_contacts(
        self,
    ) -> list[dict]:
        """
        Get distinct customers who have at least one overdue installment.
        Returns list of dicts: customer_id, email, first_name, last_name, customer_name, overdue_count, total_overdue_amount.
        """
        today = date.today()
        loans_result = await self.db.execute(
            select(Loan, Customer, Vehicle)
            .join(Customer, Loan.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
            .where(Loan.status != "closed")
        )
        # Aggregate by customer_id: overdue_count, total_amount
        customer_overdue: dict[UUID, dict] = {}
        for loan, customer, vehicle in loans_result.all():
            from_date = loan.created_at.date()
            to_date = today
            due_dates = _get_due_dates_range(loan, from_date, to_date)
            paid_per = await self._get_paid_amount_per_due_date(loan.id, completed_only=True)
            amount_per = _ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            for due_dt in due_dates:
                due_d = due_dt.date()
                if due_d >= today:
                    continue
                paid_amt = paid_per.get(due_d, 0)
                remaining = _round_amount(max(0, amount_per - paid_amt))
                if remaining <= 0:
                    continue
                amount = remaining
                cid = loan.customer_id
                if cid not in customer_overdue:
                    customer_overdue[cid] = {
                        "customer_id": cid,
                        "email": customer.email,
                        "first_name": customer.first_name,
                        "last_name": customer.last_name,
                        "customer_name": f"{customer.first_name} {customer.last_name}",
                        "overdue_count": 0,
                        "total_overdue_amount": 0.0,
                    }
                customer_overdue[cid]["overdue_count"] += 1
                customer_overdue[cid]["total_overdue_amount"] += amount
        return list(customer_overdue.values())

    async def send_bulk_overdue_reminder(
        self,
        *,
        email_subject: str | None = None,
        email_body_override: str | None = None,
        notification_title: str | None = None,
        notification_body: str | None = None,
    ) -> dict:
        """
        Send bulk email and push notification to all customers with overdue payments.
        Returns counts: customer_count, emails_sent, emails_failed, notifications_sent, no_device_count, notifications_failed.
        """
        contacts = await self.get_overdue_customer_contacts()
        customer_ids = [c["customer_id"] for c in contacts]
        emails_sent = 0
        emails_failed = 0
        for c in contacts:
            ok = await send_overdue_reminder_email(
                customer_email=c["email"],
                customer_name=c["customer_name"],
                overdue_count=c["overdue_count"],
                total_overdue_amount=c["total_overdue_amount"],
                subject=email_subject,
                body_override=email_body_override,
            )
            if ok:
                emails_sent += 1
            else:
                emails_failed += 1
        default_title = "Overdue Alert"
        default_body = "You have overdue payment(s). Please log in and pay at your earliest convenience."
        sent_count, no_device_count, failed_count = await send_notification_to_customers(
            self.db,
            title=notification_title or default_title,
            body=notification_body or default_body,
            customer_ids=customer_ids,
            data={"type": "overdue_reminder"},
        )
        return {
            "customer_count": len(contacts),
            "emails_sent": emails_sent,
            "emails_failed": emails_failed,
            "notifications_sent": sent_count,
            "no_device_count": no_device_count,
            "notifications_failed": failed_count,
        }

    async def update_payment_status_admin(
        self,
        payment_id: UUID,
        status: str,
    ) -> TransactionItem | None:
        """Admin updates payment status. Sends 'payment_confirmed' notification when status is completed."""
        payment = await self.db.get(Payment, payment_id)
        if not payment:
            return None
        payment.status = status
        self.db.add(payment)
        await self.db.commit()
        await self.db.refresh(payment)
        if status == "completed":
            await send_payment_notification(
                self.db,
                customer_id=payment.customer_id,
                notification_type="payment_confirmed",
                scope_key=scope_key_for_payment(payment.id),
                title="Payment confirmed",
                body=f"Your payment of ${_ensure_non_negative_amount(payment.amount):.2f} has been confirmed.",
            )
            await self._check_and_close_loan_if_paid(payment.loan_id)
        loan = await self.db.get(Loan, payment.loan_id)
        customer = await self.db.get(Customer, payment.customer_id)
        vehicle = await self.db.get(Vehicle, loan.vehicle_id) if loan else None
        vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
        customer_name = f"{customer.first_name} {customer.last_name}" if customer else None
        emi_val = getattr(payment, "emi_amount", None) or (loan and _ensure_non_negative_amount(loan.bi_weekly_payment_amount)) or _ensure_non_negative_amount(payment.amount)
        return TransactionItem(
            id=payment.id,
            loan_id=payment.loan_id,
            customer_id=payment.customer_id,
            customer_name=customer_name,
            vehicle_display=vehicle_display,
            amount=_ensure_non_negative_amount(payment.amount),
            emi_amount=_ensure_non_negative_amount(emi_val),
            payment_method=payment.payment_method,
            status=payment.status,
            payment_date=payment.payment_date,
            due_date=payment.due_date,
            created_at=payment.created_at,
            loan_status=_loan_status_display(loan),
        )

    async def get_receipt_for_payment(self, payment_id: UUID) -> dict | None:
        """Build receipt data for a payment. Returns None if payment not found."""
        payment = await self.db.get(Payment, payment_id)
        if not payment:
            return None
        loan = await self.db.get(Loan, payment.loan_id)
        customer = await self.db.get(Customer, payment.customer_id)
        vehicle = await self.db.get(Vehicle, loan.vehicle_id) if loan else None
        vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else None
        customer_name = f"{customer.first_name} {customer.last_name}" if customer else ""
        receipt_number = "RCP-" + str(payment.id).replace("-", "").upper()[:12]
        emi_val = getattr(payment, "emi_amount", None) or (loan and _ensure_non_negative_amount(loan.bi_weekly_payment_amount)) or _ensure_non_negative_amount(payment.amount)
        return {
            "receipt_number": receipt_number,
            "payment_id": payment.id,
            "company_name": "AutoLoanPro",
            "customer_name": customer_name,
            "customer_email": customer.email if customer else None,
            "customer_phone": customer.phone if customer else None,
            "amount": _ensure_non_negative_amount(payment.amount),
            "emi_amount": _ensure_non_negative_amount(emi_val),
            "currency": (settings.STRIPE_CURRENCY or "usd").lower(),
            "payment_method": payment.payment_method or "card",
            "payment_date": payment.payment_date,
            "due_date": payment.due_date,
            "status": payment.status,
            "loan_id": payment.loan_id,
            "vehicle_display": vehicle_display,
            "loan_status": _loan_status_display(loan),
            "note": payment.note,
            "created_at": payment.created_at,
        }

    async def get_receipt_for_customer(self, payment_id: UUID, customer_id: UUID) -> dict | None:
        """Build receipt data for a payment only if it belongs to the customer. Returns None if not found or not owned."""
        payment = await self.db.get(Payment, payment_id)
        if not payment or payment.customer_id != customer_id:
            return None
        return await self.get_receipt_for_payment(payment_id)

    def _notification_display(self, notification_type: str) -> tuple[str, str]:
        """Return (title, body) for a notification type."""
        return NOTIFICATION_TYPE_DISPLAY.get(
            notification_type,
            (notification_type.replace("_", " ").title(), "Notification sent."),
        )

    async def list_notifications_for_customer(
        self,
        customer_id: UUID,
        skip: int = 0,
        limit: int = 100,
    ) -> tuple[list[NotificationItem], int]:
        """List payment notifications for a customer. Returns (items, total)."""
        base = (
            select(PaymentNotificationLog)
            .where(PaymentNotificationLog.customer_id == customer_id)
            .order_by(PaymentNotificationLog.sent_at.desc())
        )
        total_result = await self.db.execute(
            select(func.count(PaymentNotificationLog.id)).where(
                PaymentNotificationLog.customer_id == customer_id
            )
        )
        total = total_result.scalar_one() or 0
        result = await self.db.execute(base.offset(skip).limit(limit))
        items = []
        for log in result.scalars().all():
            title, body = self._notification_display(log.notification_type)
            items.append(
                NotificationItem(
                    id=log.id,
                    notification_type=log.notification_type,
                    title=title,
                    body=body,
                    sent_at=log.sent_at,
                    customer_id=None,
                    customer_name=None,
                )
            )
        return items, total

    async def list_notifications_admin(
        self,
        customer_id: UUID | None = None,
        skip: int = 0,
        limit: int = 100,
    ) -> tuple[list[NotificationItem], int]:
        """List payment notifications for admin (all or filtered by customer_id). Returns (items, total)."""
        base = (
            select(PaymentNotificationLog, Customer)
            .join(Customer, PaymentNotificationLog.customer_id == Customer.id)
            .order_by(PaymentNotificationLog.sent_at.desc())
        )
        count_q = select(func.count(PaymentNotificationLog.id))
        if customer_id is not None:
            base = base.where(PaymentNotificationLog.customer_id == customer_id)
            count_q = count_q.where(PaymentNotificationLog.customer_id == customer_id)
        total_result = await self.db.execute(count_q)
        total = total_result.scalar_one() or 0
        result = await self.db.execute(base.offset(skip).limit(limit))
        items = []
        for log, customer in result.all():
            title, body = self._notification_display(log.notification_type)
            customer_name = f"{customer.first_name} {customer.last_name}" if customer else None
            items.append(
                NotificationItem(
                    id=log.id,
                    notification_type=log.notification_type,
                    title=title,
                    body=body,
                    sent_at=log.sent_at,
                    customer_id=log.customer_id,
                    customer_name=customer_name,
                )
            )
        return items, total
