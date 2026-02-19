import io
import uuid
import logging
from datetime import datetime, timedelta
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_

from app.api.v1.customers.schemas import _payment_schedule_description
from app.api.v1.sales.schemas import (
    CreateLeaseRequest,
    LeaseResponse,
    LeaseListItem,
    LeasesSummary,
    LeasesListResponse,
    BiWeeklyEstimateRequest,
    BiWeeklyEstimateResponse,
)
from app.core.exceptions import AppException
from app.core.utils import ensure_non_negative_amount
from app.models.customer import Customer
from app.models.vehicle import Vehicle
from app.models.loan import Loan
from app.models.customer_vehicle import CustomerVehicle
from app.models.enums import VehicleStatus, AccountStatus


def _num_payments_for_lease(term_months: int, lease_payment_type: str) -> int:
    """Number of payments over the lease term (no interest)."""
    if lease_payment_type == "monthly":
        return max(1, term_months)
    return max(1, term_months * 2)  # bi_weekly or semi_monthly


def calculate_lease_payment(
    lease_amount: float,
    down_payment: float,
    term_months: int,
    lease_payment_type: str,
) -> float:
    """Flat payment per due date (no interest)."""
    amount_financed = lease_amount - down_payment
    if amount_financed <= 0:
        return 0.0
    num_payments = _num_payments_for_lease(term_months, lease_payment_type or "bi_weekly")
    return amount_financed / num_payments


class SaleService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.logger = logging.getLogger(__name__)

    def get_bi_weekly_estimate(self, data: BiWeeklyEstimateRequest) -> BiWeeklyEstimateResponse:
        """Return estimated payment per due date without creating a lease (no interest)."""
        amount_financed = data.lease_amount - data.down_payment
        if amount_financed <= 0:
            AppException().raise_400("Down payment must be less than lease amount")
        payment_type = getattr(data, "lease_payment_type", "bi_weekly") or "bi_weekly"
        payment_amount = calculate_lease_payment(
            data.lease_amount,
            data.down_payment,
            data.term_months,
            payment_type,
        )
        return BiWeeklyEstimateResponse(
            lease_amount=data.lease_amount,
            down_payment=data.down_payment,
            amount_financed=amount_financed,
            term_months=data.term_months,
            lease_payment_type=payment_type,
            estimated_payment_amount=round(payment_amount, 2),
        )

    async def create_lease(self, data: CreateLeaseRequest) -> LeaseResponse:
        """
        Create a new vehicle lease for an existing customer.
        Associates vehicle with customer, creates loan, marks vehicle as leased.
        """
        customer = await self.db.get(Customer, data.customer_id)
        if not customer:
            AppException().raise_404("Customer not found")
        if customer.account_status != AccountStatus.active.value:
            AppException().raise_400("Customer account is inactive")

        vehicle = await self.db.get(Vehicle, data.vehicle_id)
        if not vehicle:
            AppException().raise_404("Vehicle not found")
        if vehicle.status in (VehicleStatus.sold.value, VehicleStatus.leased.value):
            AppException().raise_400("Vehicle is already leased and not available")

        existing = await self.db.execute(
            select(CustomerVehicle).where(CustomerVehicle.vehicle_id == data.vehicle_id)
        )
        if existing.scalar_one_or_none():
            AppException().raise_400("Vehicle is already leased to another customer")

        amount_financed = data.lease_amount - data.down_payment
        if amount_financed < 0:
            AppException().raise_400("Down payment cannot exceed lease amount")

        payment_type = getattr(data, "lease_payment_type", "bi_weekly") or "bi_weekly"
        bi_weekly_payment = (
            0.0
            if amount_financed <= 0
            else calculate_lease_payment(
                data.lease_amount,
                data.down_payment,
                data.term_months,
                payment_type,
            )
        )

        lease_start = datetime.utcnow()
        lease_end = lease_start + timedelta(days=round(data.term_months * 30.44)) if data.term_months > 0 else None

        cv = CustomerVehicle(
            id=uuid.uuid4(),
            customer_id=data.customer_id,
            vehicle_id=data.vehicle_id,
            lease_start_date=lease_start,
            lease_end_date=lease_end,
        )
        self.db.add(cv)
        await self.db.flush()

        vehicle.lease_price = data.lease_price
        vehicle.lease_end_date = lease_end
        vehicle.status = VehicleStatus.leased.value
        self.db.add(vehicle)
        await self.db.flush()

        loan = Loan(
            id=uuid.uuid4(),
            customer_id=data.customer_id,
            vehicle_id=data.vehicle_id,
            total_purchase_price=data.lease_amount,
            down_payment=data.down_payment,
            amount_financed=amount_financed,
            bi_weekly_payment_amount=bi_weekly_payment,
            loan_term_months=float(data.term_months),
            lease_payment_type=payment_type,
            interest_rate=None,
            status="closed" if amount_financed <= 0 else "active",
        )
        self.db.add(loan)
        await self.db.commit()
        await self.db.refresh(loan)
        await self.db.refresh(customer)
        await self.db.refresh(vehicle)

        customer_name = f"{customer.first_name} {customer.last_name}"
        vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}"
        pt = getattr(loan, "lease_payment_type", "bi_weekly") or "bi_weekly"
        payment_amt = round(ensure_non_negative_amount(loan.bi_weekly_payment_amount), 2)

        return LeaseResponse(
            loan_id=loan.id,
            customer_id=customer.id,
            customer_name=customer_name,
            vehicle_id=vehicle.id,
            vehicle_display=vehicle_display,
            lease_amount=ensure_non_negative_amount(loan.total_purchase_price),
            down_payment=ensure_non_negative_amount(loan.down_payment),
            amount_financed=ensure_non_negative_amount(loan.amount_financed),
            term_months=data.term_months,
            lease_payment_type=pt,
            bi_weekly_payment_amount=payment_amt,
            payment_amount=payment_amt,
            payment_schedule_description=_payment_schedule_description(pt),
            created_at=loan.created_at,
        )

    def _search_condition(self, term: str):
        """Condition: customer name or vehicle (make/model/year) matches search term."""
        return or_(
            Customer.first_name.ilike(term),
            Customer.last_name.ilike(term),
            Vehicle.make.ilike(term),
            Vehicle.model.ilike(term),
            Vehicle.year.ilike(term),
        )

    async def get_leases(
        self,
        customer_id: Optional[uuid.UUID] = None,
        search: Optional[str] = None,
    ) -> LeasesListResponse:
        """List leases (loans) with summary stats. Filter by customer_id and/or search (customer or vehicle)."""
        now = datetime.utcnow()
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        search_term = f"%{(search or '').strip().lower()}%" if (search and search.strip()) else None

        def apply_filters(q, *, join_customer_vehicle: bool = False):
            if join_customer_vehicle:
                q = q.join(Customer, Loan.customer_id == Customer.id).join(Vehicle, Loan.vehicle_id == Vehicle.id)
            if customer_id is not None:
                q = q.where(Loan.customer_id == customer_id)
            if search_term:
                if not join_customer_vehicle:
                    q = q.join(Customer, Loan.customer_id == Customer.id).join(Vehicle, Loan.vehicle_id == Vehicle.id)
                q = q.where(self._search_condition(search_term))
            return q

        q_summary = select(
            func.count(Loan.id).label("total_leases"),
            func.coalesce(func.sum(Loan.total_purchase_price), 0).label("total_value"),
            func.count(Loan.id).label("active_loans"),
            func.count(Loan.id).filter(Loan.created_at >= start_of_month).label("this_month"),
        ).select_from(Loan)
        q_summary = apply_filters(q_summary, join_customer_vehicle=bool(search_term))
        summary_row = (await self.db.execute(q_summary)).one()

        q = (
            select(Loan, Customer, Vehicle)
            .join(Customer, Loan.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
            .order_by(Loan.created_at.desc())
        )
        if customer_id is not None:
            q = q.where(Loan.customer_id == customer_id)
        if search_term:
            q = q.where(self._search_condition(search_term))
        result = await self.db.execute(q)
        rows = result.all()
        out = []
        for loan, customer, vehicle in rows:
            customer_name = f"{customer.first_name} {customer.last_name}"
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}"
            pt = getattr(loan, "lease_payment_type", "bi_weekly") or "bi_weekly"
            payment_amt = ensure_non_negative_amount(loan.bi_weekly_payment_amount)
            out.append(
                LeaseListItem(
                    loan_id=loan.id,
                    customer_id=loan.customer_id,
                    customer_name=customer_name,
                    vehicle_id=loan.vehicle_id,
                    vehicle_display=vehicle_display,
                    lease_amount=ensure_non_negative_amount(loan.total_purchase_price),
                    bi_weekly_payment_amount=payment_amt,
                    payment_amount=payment_amt,
                    payment_schedule_description=_payment_schedule_description(pt),
                    lease_payment_type=pt,
                    term_months=loan.loan_term_months,
                    created_at=loan.created_at,
                )
            )

        summary = LeasesSummary(
            total_leases=summary_row.total_leases or 0,
            total_value=ensure_non_negative_amount(float(summary_row.total_value or 0)),
            active_loans=summary_row.active_loans or 0,
            this_month=summary_row.this_month or 0,
        )
        return LeasesListResponse(summary=summary, leases=out)

    async def export_leases_to_excel(
        self,
        customer_id: Optional[uuid.UUID] = None,
        search: Optional[str] = None,
    ) -> bytes:
        """Export leases (loans) data to Excel (.xlsx). Same filters as get_leases."""
        search_term = f"%{(search or '').strip().lower()}%" if (search and search.strip()) else None
        q = (
            select(Loan, Customer, Vehicle)
            .join(Customer, Loan.customer_id == Customer.id)
            .join(Vehicle, Loan.vehicle_id == Vehicle.id)
            .order_by(Loan.created_at.desc())
        )
        if customer_id is not None:
            q = q.where(Loan.customer_id == customer_id)
        if search_term:
            q = q.where(self._search_condition(search_term))
        result = await self.db.execute(q)
        rows = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Leases"

        headers = [
            "Loan ID",
            "Customer ID",
            "Customer Name",
            "Vehicle ID",
            "Vehicle (Year Make Model)",
            "Lease Amount",
            "Down Payment",
            "Amount Financed",
            "Payment Amount",
            "Term (Months)",
            "Lease Payment Type",
            "Payment Schedule",
            "Loan Status",
            "Created At",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, (loan, customer, vehicle) in enumerate(rows, start=2):
            customer_name = f"{customer.first_name} {customer.last_name}"
            vehicle_display = f"{vehicle.year} {vehicle.make} {vehicle.model}"
            loan_status = getattr(loan, "status", "active")
            created_at = loan.created_at
            if hasattr(created_at, "strftime"):
                created_at_str = created_at.strftime("%Y-%m-%d %H:%M:%S")
            else:
                created_at_str = str(created_at)
            ws.cell(row=row_idx, column=1, value=str(loan.id))
            ws.cell(row=row_idx, column=2, value=str(loan.customer_id))
            ws.cell(row=row_idx, column=3, value=customer_name)
            ws.cell(row=row_idx, column=4, value=str(loan.vehicle_id))
            ws.cell(row=row_idx, column=5, value=vehicle_display)
            ws.cell(row=row_idx, column=6, value=ensure_non_negative_amount(loan.total_purchase_price))
            ws.cell(row=row_idx, column=7, value=ensure_non_negative_amount(loan.down_payment))
            ws.cell(row=row_idx, column=8, value=ensure_non_negative_amount(loan.amount_financed))
            pt = getattr(loan, "lease_payment_type", "bi_weekly") or "bi_weekly"
            ws.cell(row=row_idx, column=9, value=ensure_non_negative_amount(loan.bi_weekly_payment_amount))
            ws.cell(row=row_idx, column=10, value=loan.loan_term_months)
            ws.cell(row=row_idx, column=11, value=pt)
            ws.cell(row=row_idx, column=12, value=_payment_schedule_description(pt))
            ws.cell(row=row_idx, column=13, value=loan_status)
            ws.cell(row=row_idx, column=14, value=created_at_str)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
